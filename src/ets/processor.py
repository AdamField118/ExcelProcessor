"""
ETS Processor - Processing logic for ETS files

This processor handles ETS-specific data transformation and output generation.
It inherits from BaseProcessor and implements ETS-specific logic.
"""

import os
import sys
import shutil
import openpyxl

# Add paths for both development and frozen executable
if getattr(sys, 'frozen', False):
    base_path = sys._MEIPASS
    sys.path.insert(0, base_path)
else:
    base_path = os.path.dirname(os.path.abspath(__file__))
    sys.path.insert(0, os.path.abspath(os.path.join(base_path, '..', '..')))

try:
    from src.core.base_processor import BaseProcessor
    from src.core.file_handler import FileHandler
    from src.utils.logger import get_simple_logger
except ImportError:
    from core.base_processor import BaseProcessor
    from core.file_handler import FileHandler
    from utils.logger import get_simple_logger

logger = get_simple_logger("ets_processor")


class ETSProcessor(BaseProcessor):
    """
    ETS-specific file processor
    
    TODO: Implement ETS-specific processing logic
    """
    
    @property
    def template_path(self):
        """Path to ETS output template"""
        template_path = os.path.join("resources", "config", "ets_template.xlsx")
        
        if not os.path.exists(template_path):
            current_dir = os.path.dirname(os.path.abspath(__file__))
            project_root = os.path.dirname(os.path.dirname(current_dir))
            template_path = os.path.join(project_root, "resources", "config", "ets_template.xlsx")
        
        return template_path
    
    def process_data(self, input_data, config_values, output_path, progress_callback=None):
        """
        Process ETS files
        
        Args:
            input_data: List of DataFrames from input files
            config_values: Dictionary of configuration values
            output_path: Path where to save the output file
            progress_callback: Optional callback for progress updates
            
        Returns:
            dict: Processing results summary
        """
        try:
            if progress_callback:
                progress_callback("Setting up ETS output file...")
            
            # Copy template to output location
            if not os.path.exists(self.template_path):
                raise FileNotFoundError(f"ETS template file not found: {self.template_path}")
            
            shutil.copy2(self.template_path, output_path)
            logger.info(f"Copied template to: {output_path}")
            
            if progress_callback:
                progress_callback("Writing configuration to output...")
            
            # Write configuration values to output
            self._write_config_to_output(output_path, config_values)
            
            if progress_callback:
                progress_callback("Processing data...")
            
            # Process the data and get count
            total_records = self._process_ets_data(input_data, output_path, progress_callback)
            
            if progress_callback:
                progress_callback("Applying OCV/CCV min values and finalizing...")
            
            # Apply OCV min and CCV min to all data rows
            self._apply_min_values(output_path, config_values, total_records)
            
            # Set sheet name based on customer PO and quantity
            self._set_sheet_name(output_path, config_values, total_records)
            
            logger.info(f"ETS processing complete. Output saved to: {output_path}")
            
            return {
                'status': 'success',
                'total_files_processed': len(input_data),
                'total_records': total_records,
                'output_file': output_path
            }
            
        except Exception as e:
            error_msg = f"ETS processing failed: {str(e)}"
            if progress_callback:
                progress_callback(f"ERROR: {error_msg}")
            logger.error(error_msg)
            raise Exception(error_msg)
    
    def _apply_min_values(self, output_path, config_values, total_records):
        """
        Apply OCV min and CCV min values to all data rows
        """
        workbook = openpyxl.load_workbook(output_path)
        sheet = workbook.active
        
        try:
            ocv_min = float(config_values.get('ocv_min', '3.92'))
            ccv_min = float(config_values.get('ccv_min', '3'))
        except (ValueError, TypeError):
            ocv_min = 3.92
            ccv_min = 3
            logger.warning(f"Could not parse min values, using defaults: OCV={ocv_min}, CCV={ccv_min}")
        
        # Apply to all data rows (starting at row 8)
        for row in range(8, 8 + total_records):
            sheet.cell(row=row, column=6, value=ocv_min)  # Column F: OCV Min
            sheet.cell(row=row, column=7, value=ccv_min)  # Column G: CCV Min
        
        workbook.save(output_path)
        workbook.close()
        
        logger.info(f"Applied OCV min={ocv_min}, CCV min={ccv_min} to {total_records} rows")
    
    def _set_sheet_name(self, output_path, config_values, total_records):
        """
        Set sheet name based on customer PO and quantity
        Format: "Qty {total_records}" or "{customer_po}_Qty_{total_records}"
        """
        workbook = openpyxl.load_workbook(output_path)
        sheet = workbook.active
        
        customer_po = config_values.get('customer_po', '').strip()
        
        # Create sheet name
        if customer_po:
            sheet_name = f"Qty {total_records}"
        else:
            sheet_name = f"Qty {total_records}"
        
        # Ensure sheet name is valid (Excel has 31 char limit)
        if len(sheet_name) > 31:
            sheet_name = sheet_name[:31]
        
        sheet.title = sheet_name
        
        workbook.save(output_path)
        workbook.close()
        
        logger.info(f"Set sheet name to: {sheet_name}")
    
    def _write_config_to_output(self, output_path, config_values):
        """
        Write configuration values to the output file header
        Writes to cells based on ETS template structure
        """
        workbook = openpyxl.load_workbook(output_path)
        sheet = workbook.active
        
        # Write configuration to header area
        sheet['C2'] = config_values.get('part_name', '')
        sheet['C3'] = config_values.get('revision_number', '')
        sheet['C4'] = config_values.get('customer_pn', '')
        sheet['C5'] = config_values.get('customer_po', '')
        
        workbook.save(output_path)
        workbook.close()
        
        logger.info("Configuration written to output file")
    
    def _process_ets_data(self, input_data, output_path, progress_callback=None):
        """
        Main ETS data processing logic
        
        Extracts data from input files where column C (Pass Fail) = TRUE
        and writes to output file with proper formatting.
        
        Args:
            input_data: List of DataFrames
            output_path: Output file path
            progress_callback: Progress callback function
            
        Returns:
            int: Total number of records processed
        """
        total_records = 0
        all_parts = []
        
        # Extract parts from all input files
        for df_idx, df in enumerate(input_data):
            if progress_callback:
                progress_callback(f"Processing file {df_idx + 1}/{len(input_data)}...")
            
            logger.info(f"Processing DataFrame {df_idx + 1}, shape: {df.shape}")
            
            # Find rows where column C (index 2, "Pass Fail") = True
            # Skip header rows by starting from a reasonable index
            # The data typically starts after row 2 (pandas will have header info)
            
            for row_idx in range(len(df)):
                try:
                    # Check if column C (index 2) is True
                    pass_fail_value = df.iloc[row_idx, 2]
                    
                    # Handle both boolean True and string "True"
                    if pass_fail_value == True or pass_fail_value == 'True':
                        # Extract values from this row
                        part_data = {
                            'part_number': df.iloc[row_idx, 0],      # Column A (index 0)
                            'cell_date_code': df.iloc[row_idx, 1],   # Column B (index 1)
                            'pass_fail': True,                        # Column C (always True for extracted rows)
                            'ocv_value': df.iloc[row_idx, 5],        # Column F (index 5)
                            'ccv_value': df.iloc[row_idx, 6],        # Column G (index 6)
                            'test_date_time': df.iloc[row_idx, 11],  # Column L (index 11)
                        }
                        
                        all_parts.append(part_data)
                        total_records += 1
                        
                except (IndexError, KeyError) as e:
                    # Skip rows that don't have all required columns
                    logger.debug(f"Skipping row {row_idx} in file {df_idx + 1}: {e}")
                    continue
            
            if progress_callback:
                progress = int(((df_idx + 1) / len(input_data)) * 50)  # First 50% for extraction
                progress_callback(f"Extracted {total_records} parts from {df_idx + 1}/{len(input_data)} files ({progress}%)")
        
        logger.info(f"Extracted {total_records} total parts from all files")
        
        # Now write all parts to output file
        if progress_callback:
            progress_callback(f"Writing {total_records} parts to output file...")
        
        workbook = openpyxl.load_workbook(output_path)
        sheet = workbook.active
        
        # Get OCV min and CCV min from the file (should already be in sheet)
        # These will be read from config and applied to all rows
        # We'll set them when writing rows
        
        # Starting row for data (row 8 in Excel, 0-indexed is row 7)
        current_row = 8
        
        for idx, part in enumerate(all_parts):
            # Write data to columns A through H
            sheet.cell(row=current_row, column=1, value=part['part_number'])       # A: Part Number
            sheet.cell(row=current_row, column=2, value=part['cell_date_code'])    # B: Cell Date Code
            sheet.cell(row=current_row, column=3, value='True')                    # C: Pass Fail (always True)
            sheet.cell(row=current_row, column=4, value=part['ocv_value'])         # D: OCV Value
            sheet.cell(row=current_row, column=5, value=part['ccv_value'])         # E: CCV Value
            # Columns F and G (OCV min, CCV min) will be set from config values
            sheet.cell(row=current_row, column=8, value=part['test_date_time'])    # H: Test Date Time
            
            current_row += 1
            
            # Update progress every 100 parts
            if idx % 100 == 0 and progress_callback:
                progress = 50 + int((idx / total_records) * 50)  # Second 50% for writing
                progress_callback(f"Writing parts to output: {idx}/{total_records} ({progress}%)")
        
        workbook.save(output_path)
        workbook.close()
        
        logger.info(f"Successfully wrote {total_records} parts to output file")
        return total_records
    
    @staticmethod
    def validate_config(config_values):
        """
        Validate ETS configuration values
        All fields are required
        """
        required_fields = ['part_name', 'revision_number', 'customer_pn', 'customer_po', 'ocv_min', 'ccv_min']
        errors = []
        
        for field in required_fields:
            value = config_values.get(field, '')
            if not isinstance(value, str) or not value.strip():
                # Make field names more readable
                field_display = field.replace('_', ' ').title()
                errors.append(f"{field_display} is required")
        
        # Validate that OCV min and CCV min are numeric
        try:
            ocv_min = config_values.get('ocv_min', '')
            if ocv_min.strip():
                float(ocv_min)
        except (ValueError, AttributeError):
            errors.append("OCV Min must be a valid number")
        
        try:
            ccv_min = config_values.get('ccv_min', '')
            if ccv_min.strip():
                float(ccv_min)
        except (ValueError, AttributeError):
            errors.append("CCV Min must be a valid number")
        
        return errors